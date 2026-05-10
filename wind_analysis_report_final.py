"""
wind_analysis_report_final.py
==============================
NE6010 Offshore Wind Engineering – Assignment Analysis Script

Purpose
-------
Performs a two-question wind and wave resource assessment for an offshore
wind site located off the west coast of Ireland, using observational data
from the Irish Weather Buoy Network M2 buoy and SWAN model wave hindcast.

Questions covered
-----------------
  Q1 – Wind Resource Assessment
       · Hub-height extrapolation (power law, IEC 61400-3)
       · Weibull distribution fitting (MLE)
       · Wind-speed statistics, histograms, time-series, wind rose, vertical profile

  Q2 – Wave Climate & Operational Accessibility
       · Significant wave height (Hs) statistics and Weibull fit
       · Vessel operational accessibility (CTV and Jack-up Barge)
         – combined Hs + wind limits, consecutive-hour qualifying blocks
         – all-hours and daylight-hours-only analyses
       · Weibull-based wind operability estimation

Data sources
------------
  · IWBNetwork_M2_2024_detailed.xlsx  – hourly buoy wind/wave data for 2024
  · IWBNetwork_5_years.csv            – 5-year M2 buoy wind record (2021-2025)
  · wave_data_swan.xlsx               – SWAN hindcast Hs, hourly, 2021-2025

Turbine assumed
---------------
  Generic 15 MW offshore turbine: rotor diameter 220 m, hub height 150 m
  (cf. IEA 15 MW reference turbine: D = 240 m, H = 150 m)

Outputs
-------
  Question 1 results/   – Excel summary + PNG plots
  Question 2 output/    – Excel summary + PNG plots

Author : NE6010 Student 125135282 - Aoife Prendergast
Date   : 2026
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime
from scipy.stats import weibull_min

# Load the Excel file
file_path = 'IWBNetwork_M2_2024_detailed.xlsx'

# Read the file, skipping the units row
df = pd.read_excel(file_path, header=0)

# Remove the units row (second row)
df = df.iloc[1:].reset_index(drop=True)

# Question 1: Wind Resource Assessment

# Convert data types
df['longitude'] = pd.to_numeric(df['longitude'], errors='coerce')
df['latitude'] = pd.to_numeric(df['latitude'], errors='coerce')
df['WindSpeed'] = pd.to_numeric(df['WindSpeed'], errors='coerce')
df['WaveHeight'] = pd.to_numeric(df['WaveHeight'], errors='coerce')
df['WindDirection'] = pd.to_numeric(df['WindDirection'], errors='coerce')
df['time'] = pd.to_datetime(df['time'], errors='coerce')
df['Hmax'] = pd.to_numeric(df['Hmax'], errors='coerce')

# Convert wind speed from knots to m/s (1 knot = 0.514444 m/s)
df['WindSpeed_ms'] = df['WindSpeed'] * 0.514444

# ── Hub-height extrapolation (power law) ─────────────────────
# Rotor diameter D = 220 m; lower blade tip clearance ≥ 22 m (IEC 61400-3 offshore)
#   → H_hub ≥ 22 + 110 = 132 m  →  adopt H_hub = 150 m
#   (cf. IEA 15 MW reference turbine: D = 240 m, H = 150 m)
# M2 buoy anemometer: 4 m above MSL (Guesstimate based on typical buoy designs)
# Power law: V(H) = V_ref × (H / H_ref)^α,  α = 0.11 (IEC 61400-3 offshore)
ROTOR_DIAMETER   = 220
HUB_HEIGHT       = 150       # m above MSL
REF_HEIGHT       = 4         # m – M2 buoy anemometer
WIND_SHEAR_ALPHA = 0.11      # offshore (IEC 61400-3)
hub_factor = (HUB_HEIGHT / REF_HEIGHT) ** WIND_SHEAR_ALPHA
df['WindSpeed_hub'] = df['WindSpeed_ms'] * hub_factor

# Add hour and date columns for analysis
df['hour'] = df['time'].dt.hour
df['date'] = df['time'].dt.date

# Display basic info to console
print("="*60)
print("WIND DATA ANALYSIS")
print("="*60)
print(f"\nDataset Overview:")
print(f"Total records: {len(df)}")
print(f"Date range: {df['time'].min()} to {df['time'].max()}")
print(f"Number of days analyzed: {df['date'].nunique()}")
print(f"Total hours: {len(df)}")

# Check for missing days
import datetime
start_date = df['time'].min().date()
end_date = df['time'].max().date()
all_days = set(start_date + datetime.timedelta(n) for n in range((end_date - start_date).days + 1))
present_days = set(df['date'].unique())
missing_days = sorted(all_days - present_days)
if missing_days:
    print(f"\nMissing days ({len(missing_days)}): {missing_days}")
else:
    print("\nNo missing days in dataset.")

# Site coordinates (M2 buoy)
M2_LAT = 53.4836
M2_LON = -5.4302

SITE2_LAT = 53.6016963
SITE2_LON = -5.8826392
#

# ==============================================================
# QUESTION 1 - WIND RESOURCE ASSESSMENT
# ==============================================================

# ── Hub-height justification ───────────────────────────────────
print("\n" + "="*60)
print("HUB HEIGHT SELECTION")
print("="*60)
print(f"\n  Rotor diameter            : {ROTOR_DIAMETER} m")
print(f"  Blade radius              : {ROTOR_DIAMETER/2:.0f} m")
print(f"  Min blade tip clearance   : 22 m (IEC 61400-3 offshore)")
print(f"  Min hub height            : {ROTOR_DIAMETER/2 + 22:.0f} m")
print(f"  Adopted hub height        : {HUB_HEIGHT} m")
print(f"  (cf. IEA 15 MW ref turbine: D=240 m, H=150 m)")
print(f"\n  Power-law shear exponent a: {WIND_SHEAR_ALPHA} (IEC 61400-3 offshore)")
print(f"  Reference height          : {REF_HEIGHT} m (M2 buoy anemometer)")
print(f"  Extrapolation factor      : ({HUB_HEIGHT}/{REF_HEIGHT})^{WIND_SHEAR_ALPHA} = {hub_factor:.4f}")

# ── Wind statistics ────────────────────────────────────────────
print("\n" + "="*60)
print("WIND RESOURCE STATISTICS - ALL HOURS")
print("="*60)

print(f"\nAt measurement height ({REF_HEIGHT} m) [m/s]:")
print(f"  Mean    : {df['WindSpeed_ms'].mean():.2f}")
print(f"  Median  : {df['WindSpeed_ms'].median():.2f}")
print(f"  Std Dev : {df['WindSpeed_ms'].std():.2f}")
print(f"  Min     : {df['WindSpeed_ms'].min():.2f}")
print(f"  Max     : {df['WindSpeed_ms'].max():.2f}")

print(f"\nAt hub height ({HUB_HEIGHT} m) [m/s]:")
print(f"  Mean    : {df['WindSpeed_hub'].mean():.2f}")
print(f"  Median  : {df['WindSpeed_hub'].median():.2f}")
print(f"  Std Dev : {df['WindSpeed_hub'].std():.2f}")
print(f"  Min     : {df['WindSpeed_hub'].min():.2f}")
print(f"  Max     : {df['WindSpeed_hub'].max():.2f}")

# ── Wind speed distribution at hub height ─────────────────────
print("\n" + "="*60)
print(f"WIND SPEED DISTRIBUTION AT HUB HEIGHT ({HUB_HEIGHT} m)")
print("="*60)
bins_hub   = [0, 7, 10, 13, 16, 20, 60]
labels_hub = ['0-7', '7-10', '10-13', '13-16', '16-20', '20+']
df['wind_cat_hub'] = pd.cut(df['WindSpeed_hub'], bins=bins_hub, labels=labels_hub)
print(df['wind_cat_hub'].value_counts().sort_index())

# ── Weibull fits ───────────────────────────────────────────────
# Maximum-likelihood estimation (MLE) is used.  floc=0 fixes the location
# parameter to zero, which is physically correct: wind speed is bounded
# below by 0 m/s and the two-parameter Weibull (shape k, scale C) is the
# standard form used in IEC 61400-1 / IEC 61400-3 resource assessments.
wind_data     = df['WindSpeed_ms'].dropna()
wind_data_hub = df['WindSpeed_hub'].dropna()
shape_k,     loc,     scale_c     = weibull_min.fit(wind_data,     floc=0)
shape_k_hub, loc_hub, scale_c_hub = weibull_min.fit(wind_data_hub, floc=0)

print("\n" + "="*60)
print("WEIBULL DISTRIBUTION PARAMETERS")
print("="*60)
print(f"\nAt measurement height ({REF_HEIGHT} m):")
print(f"  Shape parameter k : {shape_k:.4f}")
print(f"  Scale parameter C : {scale_c:.4f} m/s")
print(f"\nAt hub height ({HUB_HEIGHT} m):")
print(f"  Shape parameter k : {shape_k_hub:.4f}")
print(f"  Scale parameter C : {scale_c_hub:.4f} m/s")

# ──────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────────────────────
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_DIR = 'Question 1 results'
os.makedirs(OUT_DIR, exist_ok=True)

wb = Workbook()

# ── Helper ────────────────────────────────────────────────────
HDR_FONT  = Font(bold=True, color='FFFFFF')
HDR_FILL  = PatternFill('solid', fgColor='1F3864')
SUB_FILL  = PatternFill('solid', fgColor='D6E4F0')
SUB_FONT  = Font(bold=True)

def write_header(ws, row, text, ncols=2):
    """Write a dark-blue section header spanning `ncols` columns in worksheet `ws`."""
    ws.cell(row=row, column=1, value=text).font  = HDR_FONT
    ws.cell(row=row, column=1).fill              = HDR_FILL
    ws.cell(row=row, column=1).alignment         = Alignment(horizontal='left')
    for c in range(2, ncols + 1):
        ws.cell(row=row, column=c).fill = HDR_FILL

def write_subheader(ws, row, text):
    """Write a light-blue sub-header row across columns 1-2 in worksheet `ws`."""
    ws.cell(row=row, column=1, value=text).font = SUB_FONT
    ws.cell(row=row, column=1).fill             = SUB_FILL
    ws.cell(row=row, column=2).fill             = SUB_FILL

def write_row(ws, row, label, value):
    """Write a label-value pair into columns 1 and 2 of `ws` at the given `row`."""
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=value)

def autofit(ws, col_widths):
    """Set explicit column widths (characters) for worksheet `ws` using the provided list."""
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

# ── Sheet 1: Dataset Overview ─────────────────────────────────
ws1 = wb.active
ws1.title = 'Dataset Overview'
r = 1
write_header(ws1, r, 'DATASET OVERVIEW'); r += 1
ws1.cell(row=r, column=1, value='Parameter'); ws1.cell(row=r, column=2, value='Value')
ws1.cell(row=r, column=1).font = SUB_FONT; ws1.cell(row=r, column=2).font = SUB_FONT
r += 1
write_row(ws1, r, 'Total records',          len(df));                                    r += 1
write_row(ws1, r, 'Date range start',        str(df['time'].min()));                     r += 1
write_row(ws1, r, 'Date range end',          str(df['time'].max()));                     r += 1
write_row(ws1, r, 'Days analysed',           df['date'].nunique());                      r += 1
write_row(ws1, r, 'Missing days (count)',    len(missing_days));                         r += 1
write_row(ws1, r, 'Missing dates',           ', '.join(str(d) for d in missing_days));  r += 1
r += 1
write_header(ws1, r, 'HUB HEIGHT SELECTION'); r += 1
write_row(ws1, r, 'Rotor diameter (m)',              ROTOR_DIAMETER);                    r += 1
write_row(ws1, r, 'Blade radius (m)',                ROTOR_DIAMETER / 2);               r += 1
write_row(ws1, r, 'Min blade tip clearance (m)',     22);                                r += 1
write_row(ws1, r, 'Min hub height (m)',              ROTOR_DIAMETER / 2 + 22);          r += 1
write_row(ws1, r, 'Adopted hub height (m)',          HUB_HEIGHT);                       r += 1
write_row(ws1, r, 'Power-law shear exponent α',      WIND_SHEAR_ALPHA);                 r += 1
write_row(ws1, r, 'Reference height (m)',            REF_HEIGHT);                       r += 1
write_row(ws1, r, 'Extrapolation factor',            round(hub_factor, 4));             r += 1
autofit(ws1, [38, 40])

# ── Sheet 2: Wind Statistics ──────────────────────────────────
ws2 = wb.create_sheet('Wind Statistics')
r = 1
write_header(ws2, r, 'WIND RESOURCE STATISTICS - ALL HOURS', ncols=3); r += 1
ws2.cell(row=r, column=1, value='Statistic')
ws2.cell(row=r, column=2, value=f'At {REF_HEIGHT} m measurement height (m/s)')
ws2.cell(row=r, column=3, value=f'At {HUB_HEIGHT} m hub height (m/s)')
for c in range(1, 4):
    ws2.cell(row=r, column=c).font = SUB_FONT
r += 1
for stat, v_ref, v_hub in [
    ('Mean',    round(df['WindSpeed_ms'].mean(),   2), round(df['WindSpeed_hub'].mean(),   2)),
    ('Median',  round(df['WindSpeed_ms'].median(), 2), round(df['WindSpeed_hub'].median(), 2)),
    ('Std Dev', round(df['WindSpeed_ms'].std(),    2), round(df['WindSpeed_hub'].std(),    2)),
    ('Min',     round(df['WindSpeed_ms'].min(),    2), round(df['WindSpeed_hub'].min(),    2)),
    ('Max',     round(df['WindSpeed_ms'].max(),    2), round(df['WindSpeed_hub'].max(),    2)),
]:
    ws2.cell(row=r, column=1, value=stat)
    ws2.cell(row=r, column=2, value=v_ref)
    ws2.cell(row=r, column=3, value=v_hub)
    r += 1
autofit(ws2, [12, 38, 30])

# ── Sheet 3: Wind Speed Distribution ─────────────────────────
ws3 = wb.create_sheet('Wind Distribution')
r = 1
write_header(ws3, r, f'WIND SPEED DISTRIBUTION AT HUB HEIGHT ({HUB_HEIGHT} m)', ncols=3); r += 1
ws3.cell(row=r, column=1, value='Range (m/s)')
ws3.cell(row=r, column=2, value='Count (hours)')
ws3.cell(row=r, column=3, value='Frequency (%)')
for c in range(1, 4):
    ws3.cell(row=r, column=c).font = SUB_FONT
r += 1
dist = df['wind_cat_hub'].value_counts().sort_index()
total = dist.sum()
for label, count in dist.items():
    ws3.cell(row=r, column=1, value=str(label))
    ws3.cell(row=r, column=2, value=int(count))
    ws3.cell(row=r, column=3, value=round(count / total * 100, 2))
    r += 1
ws3.cell(row=r, column=1, value='Total').font = SUB_FONT
ws3.cell(row=r, column=2, value=int(total)).font = SUB_FONT
ws3.cell(row=r, column=3, value=100.0).font = SUB_FONT
autofit(ws3, [16, 18, 16])

# ── Sheet 4: Weibull Parameters ───────────────────────────────
ws4 = wb.create_sheet('Weibull Parameters')
r = 1
write_header(ws4, r, 'WEIBULL DISTRIBUTION PARAMETERS', ncols=3); r += 1
ws4.cell(row=r, column=1, value='Parameter')
ws4.cell(row=r, column=2, value=f'At {REF_HEIGHT} m measurement height')
ws4.cell(row=r, column=3, value=f'At {HUB_HEIGHT} m hub height')
for c in range(1, 4):
    ws4.cell(row=r, column=c).font = SUB_FONT
r += 1
ws4.cell(row=r, column=1, value='Shape parameter k')
ws4.cell(row=r, column=2, value=round(shape_k,     4))
ws4.cell(row=r, column=3, value=round(shape_k_hub, 4))
r += 1
ws4.cell(row=r, column=1, value='Scale parameter C (m/s)')
ws4.cell(row=r, column=2, value=round(scale_c,     4))
ws4.cell(row=r, column=3, value=round(scale_c_hub, 4))
autofit(ws4, [26, 30, 26])

excel_path = os.path.join(OUT_DIR, 'Q1_wind_analysis_summary.xlsx')
try:
    wb.save(excel_path)
    print(f"\nExcel summary saved: {excel_path}")
except PermissionError:
    import datetime as _dt
    excel_path = os.path.join(OUT_DIR, f'Q1_wind_analysis_summary_{_dt.datetime.now().strftime("%H%M%S")}.xlsx')
    wb.save(excel_path)
    print(f"\nExcel summary saved (original file was open): {excel_path}")

# ──────────────────────────────────────────────────────────────
# PLOTS
# ──────────────────────────────────────────────────────────────

print("\n" + "="*60)
print("GENERATING PLOTS")
print("="*60)

# Plot 1: Wind speed time series – buoy vs hub height
fig, axes = plt.subplots(2, 1, figsize=(12, 8), sharex=True)
axes[0].plot(df['time'], df['WindSpeed_ms'], linewidth=0.4, color='steelblue')
axes[0].axhline(df['WindSpeed_ms'].mean(), color='red', linestyle='--', linewidth=1.2,
                label=f"Mean = {df['WindSpeed_ms'].mean():.2f} m/s")
axes[0].set_ylabel(f'Wind Speed at {REF_HEIGHT} m (m/s)', fontsize=11)
axes[0].set_title('Wind Speed Time Series', fontsize=13, fontweight='bold')
axes[0].legend(fontsize=9)
axes[0].grid(True, alpha=0.3)

axes[1].plot(df['time'], df['WindSpeed_hub'], linewidth=0.4, color='darkorange')
axes[1].axhline(df['WindSpeed_hub'].mean(), color='red', linestyle='--', linewidth=1.2,
                label=f"Mean = {df['WindSpeed_hub'].mean():.2f} m/s")
axes[1].set_ylabel(f'Wind Speed at {HUB_HEIGHT} m hub (m/s)', fontsize=11)
axes[1].set_xlabel('Time', fontsize=11)
axes[1].legend(fontsize=9)
axes[1].grid(True, alpha=0.3)
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig(os.path.join(OUT_DIR, 'Q1_wind_timeseries.png'), dpi=300, bbox_inches='tight')
print("Saved: Q1_wind_timeseries.png")
plt.close()

# Plot 2: Side-by-side histogram + Weibull at both heights
fig, axes = plt.subplots(1, 2, figsize=(14, 6))
for ax, data, k, C, color, title in [
    (axes[0], wind_data,     shape_k,     scale_c,     'steelblue',  f'Measurement Height ({REF_HEIGHT} m)'),
    (axes[1], wind_data_hub, shape_k_hub, scale_c_hub, 'darkorange', f'Hub Height ({HUB_HEIGHT} m)'),
]:
    ax.hist(data, bins=30, density=True, edgecolor='black', alpha=0.65,
            color=color, label='Observed')
    x = np.linspace(0, data.max(), 1000)
    ax.plot(x, weibull_min.pdf(x, k, loc=0, scale=C), 'r-', linewidth=2,
            label=f'Weibull\nk={k:.2f}, C={C:.2f} m/s')
    ax.axvline(data.mean(), color='navy', linestyle='--', linewidth=1.5,
               label=f'Mean = {data.mean():.2f} m/s')
    ax.set_title(title, fontsize=12, fontweight='bold')
    ax.set_xlabel('Wind Speed (m/s)', fontsize=11)
    ax.set_ylabel('Probability Density', fontsize=11)
    ax.legend(fontsize=9)
    ax.grid(True, alpha=0.3)
plt.suptitle('Wind Speed Frequency Distribution & Weibull Fit', fontsize=13, fontweight='bold')
plt.tight_layout()
plt.savefig(os.path.join(OUT_DIR, 'Q1_wind_histogram_weibull.png'), dpi=300, bbox_inches='tight')
print("Saved: Q1_wind_histogram_weibull.png")
plt.close()

# Plot 3: Vertical wind profile (power law)
fig, ax = plt.subplots(figsize=(6, 8))
heights = np.linspace(0.5, HUB_HEIGHT + ROTOR_DIAMETER / 2 + 10, 400)
v_profile = df['WindSpeed_ms'].mean() * (heights / REF_HEIGHT) ** WIND_SHEAR_ALPHA
ax.plot(v_profile, heights, 'b-', linewidth=2, label='Power-law profile')
ax.axhline(REF_HEIGHT, color='steelblue', linestyle='--', linewidth=1.5,
           label=f'Buoy ({REF_HEIGHT} m): {df["WindSpeed_ms"].mean():.2f} m/s')
ax.axhline(HUB_HEIGHT, color='darkorange', linestyle='--', linewidth=2,
           label=f'Hub ({HUB_HEIGHT} m): {df["WindSpeed_hub"].mean():.2f} m/s')
ax.axhspan(HUB_HEIGHT - ROTOR_DIAMETER / 2, HUB_HEIGHT + ROTOR_DIAMETER / 2,
           alpha=0.12, color='green',
           label=f'Rotor swept ({HUB_HEIGHT - ROTOR_DIAMETER/2:.0f}–{HUB_HEIGHT + ROTOR_DIAMETER/2:.0f} m)')
ax.set_xlabel('Mean Wind Speed (m/s)', fontsize=12)
ax.set_ylabel('Height above MSL (m)', fontsize=12)
ax.set_title(f'Vertical Wind Speed Profile\n(Power Law, \u03b1={WIND_SHEAR_ALPHA})',
             fontsize=13, fontweight='bold')
ax.legend(fontsize=9)
ax.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(OUT_DIR, 'Q1_wind_profile.png'), dpi=300, bbox_inches='tight')
print("Saved: Q1_wind_profile.png")
plt.close()

# Plot 5: Weibull CDF at hub height
fig, ax = plt.subplots(figsize=(10, 6))
x_w = np.linspace(0, wind_data_hub.max(), 500)
cdf_hub = 1 - np.exp(-(x_w / scale_c_hub) ** shape_k_hub)
ax.plot(x_w, cdf_hub * 100, color='darkorange', linewidth=2,
        label=f'Weibull CDF  k={shape_k_hub:.2f}, C={scale_c_hub:.2f} m/s')
sorted_hub = np.sort(wind_data_hub.values)
ecdf_hub = np.arange(1, len(sorted_hub) + 1) / len(sorted_hub) * 100
ax.plot(sorted_hub, ecdf_hub, 'grey', linewidth=0.8, alpha=0.6, label='Empirical CDF')
ax.set_xlabel(f'Wind Speed at {HUB_HEIGHT} m Hub Height (m/s)', fontsize=12)
ax.set_ylabel('Cumulative Probability (%)', fontsize=12)
ax.set_title(f'Weibull CDF at Hub Height ({HUB_HEIGHT} m)', fontsize=13, fontweight='bold')
ax.legend(fontsize=10)
ax.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(OUT_DIR, 'Q1_weibull_cdf_hub.png'), dpi=300, bbox_inches='tight')
print("Saved: Q1_weibull_cdf_hub.png")
plt.close()

# Wind Rose
from windrose import WindroseAxes

wr_dir = df['WindDirection'].dropna()
wr_spd = df['WindSpeed_ms'].loc[wr_dir.index]

fig = plt.figure(figsize=(8, 8))
ax_wr = WindroseAxes.from_ax(fig=fig)
ax_wr.bar(
    wr_dir, wr_spd,
    normed=True,
    bins=[0, 3, 6, 9, 12, 16, 24],
    colors=['#ffffcc', '#a1dab4', '#41b6c4', '#2c7fb8', '#253494', '#7b2d8b', '#4d0026'],
    opening=0.9, edgecolor='white', linewidth=0.3
)
ax_wr.set_legend(
    title=f'Wind Speed at {REF_HEIGHT} m buoy (m/s)',
    loc='lower right', bbox_to_anchor=(1.15, -0.12), fontsize=9
)
ax_wr.set_title(f'Wind Rose \u2013 M2 Buoy (2024)\n(frequency by direction, coloured by buoy wind speed at {REF_HEIGHT} m)',
                fontsize=12, fontweight='bold', pad=20)
import warnings
with warnings.catch_warnings():
    warnings.simplefilter('ignore')
    plt.tight_layout()
plt.savefig(os.path.join(OUT_DIR, f'Q1_wind_rose_buoy_{REF_HEIGHT}m.png'), dpi=300, bbox_inches='tight')
print(f"Saved: Q1_wind_rose_buoy_{REF_HEIGHT}m.png")
plt.close()

print("\n" + "="*60)
print("Question 1 - ANALYSIS COMPLETE")
print("="*60)

# ==============================================================
# QUESTION 2 – Wave Climate & Operational Accessibility
# ==============================================================

import calendar
import matplotlib.ticker as mticker
import matplotlib.patches as mpatches
from matplotlib.colors import ListedColormap

# ── Output directory ──────────────────────────────────────────
Q2_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Question 2 output')
os.makedirs(Q2_DIR, exist_ok=True)

# ── Round-trip window calculation (used in both Q2 and Q2 extra) ──────────
import math

DUBLIN_PORT_LAT =  53.3498   # Alexandra Wharf
DUBLIN_PORT_LON =  -6.2603

def _haversine_nm(lat1, lon1, lat2, lon2):
    R_km  = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi  = math.radians(lat2 - lat1)
    dlam  = math.radians(lon2 - lon1)
    a     = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return 2 * R_km * math.asin(math.sqrt(a)) / 1.852   # nautical miles

SITE_DIST_NM = _haversine_nm(DUBLIN_PORT_LAT, DUBLIN_PORT_LON, SITE2_LAT, SITE2_LON)  # [NM] – reserved for Q3 vessel transit calculations
SITE_DIST_KM = SITE_DIST_NM * 1.852                                                    # [km]
M2_DIST_NM   = _haversine_nm(DUBLIN_PORT_LAT, DUBLIN_PORT_LON, M2_LAT,    M2_LON)    # [NM] – reserved for Q3 vessel transit calculations
M2_DIST_KM   = M2_DIST_NM * 1.852                                                      # [km]

# CTV:  25 knots transit speed; 2 h crew/material transfer window (no on-site stay)
# JUB:  6 knots transit speed; 72 h on-site (jack-up, campaign work, jack-down)
VESSEL_PARAMS = {
    'Crew Transfer Vessel': {
        'transit_kts':  25.0,
        'onsite_h':      2.0,
        'max_wind_ms':  12.0,
        'max_hs':        0.9,
        'color':        '#2196F3',
    },
    'Jack-up Barge': {
        'transit_kts':   6.0,
        'onsite_h':     72.0,
        'max_wind_ms':  10.0,
        'max_hs':        2.5,
        'color':        '#FF9800',
    },
}

for _vessel, _vp in VESSEL_PARAMS.items():
    _leg_h            = SITE_DIST_NM / _vp['transit_kts']
    _rt_h_exact       = 2 * _leg_h + _vp['onsite_h']
    _vp['one_way_h']  = _leg_h
    _vp['rt_h_exact'] = _rt_h_exact
    _vp['window_h']   = math.ceil(_rt_h_exact)

print("\n" + "="*60)
print("QUESTION 2 – Wave Climate Assessment")
print("="*60)

# ── Load wave data ────────────────────────────────────────────
wave_df = pd.read_excel('wave_data_swan.xlsx', sheet_name='Data', header=0)
wave_df['time'] = pd.to_datetime(wave_df['time'])
wave_df = wave_df.sort_values('time').reset_index(drop=True)
wave_df['hs'] = pd.to_numeric(wave_df['hs'], errors='coerce')
wave_df['date'] = wave_df['time'].dt.date
wave_df['year'] = wave_df['time'].dt.year
wave_df['month'] = wave_df['time'].dt.month

print(f"Wave data loaded: {wave_df['time'].min().date()} to {wave_df['time'].max().date()}")
print(f"Total hourly records: {len(wave_df):,}")

# ── Mean Hs ───────────────────────────────────────────────────
mean_hs    = wave_df['hs'].mean()             # [m]
median_hs  = wave_df['hs'].median()           # [m]
std_hs     = wave_df['hs'].std()              # [m]
p90_hs     = wave_df['hs'].quantile(0.90)     # [m] 90th-percentile Hs
max_hs_obs = wave_df['hs'].max()              # [m] maximum observed Hs

print(f"\nSite Mean Significant Wave Height (Hs): {mean_hs:.3f} m")
print(f"  Median Hs:   {median_hs:.3f} m")
print(f"  Std Dev Hs:  {std_hs:.3f} m")
print(f"  90th pctile: {p90_hs:.3f} m")
print(f"  Maximum Hs:  {max_hs_obs:.3f} m")

# ── Vessel definitions (wave + wind limits) ─────────────────
# Operational limits sourced from assignment brief and typical industry values:
#   CTV  – Hs ≤ 0.9 m (DNV-ST-0437 / typical North Sea O&M practice),
#           wind ≤ 12 m/s (Beaufort 6 – upper safe limit for personnel transfer)
#   Jack-up Barge – Hs ≤ 2.5 m, wind ≤ 10 m/s (typical jack-up weather criteria)
VESSELS = {
    'Crew Transfer Vessel': {'max_hs': 0.9, 'max_wind_ms': 12.0, 'color': '#2196F3'},
    'Jack-up Barge':        {'max_hs': 2.5, 'max_wind_ms': 10.0, 'color': '#FF9800'},
}

# Minimum qualifying block length: 12 consecutive hours for both vessels
MIN_WINDOW_HOURS = 4

# ── Load 5-year M2 wind data and merge hourly with wave_df ──
_wind5 = pd.read_csv('IWBNetwork_5_years.csv', skiprows=[1])
_wind5['time']     = pd.to_datetime(_wind5['time'], utc=True).dt.tz_localize(None)
_wind5['wind_ms']  = pd.to_numeric(_wind5['WindSpeed'], errors='coerce') * 0.514444
_wind5 = _wind5[['time', 'wind_ms']]
wave_df = wave_df.merge(_wind5, on='time', how='left')
print(f"Wind data merged: {_wind5['time'].min().date()} to {_wind5['time'].max().date()}")
print(f"  Records with wind obs : {wave_df['wind_ms'].notna().sum():,} / {len(wave_df):,}")

# ── Sunrise / sunset for each day at the site (UTC) ──────────
from astral import LocationInfo as _LocInfo
from astral.sun import sun as _sun
from zoneinfo import ZoneInfo as _ZI

_site_obs = _LocInfo(name='OffshoreWindSite', region='IE',
                     timezone='UTC',
                     latitude=SITE2_LAT, longitude=SITE2_LON).observer
_utc_zi   = _ZI('UTC')

unique_dates = sorted(wave_df['date'].unique())
_ss_rows = []
for _d in unique_dates:
    _s = _sun(_site_obs, date=_d, tzinfo=_utc_zi)
    _ss_rows.append({
        'date':        _d,
        'sunrise_utc': _s['sunrise'].replace(tzinfo=None),
        'sunset_utc':  _s['sunset'].replace(tzinfo=None),
    })
sunrise_sunset_df = pd.DataFrame(_ss_rows)
sunrise_sunset_df['daylight_h'] = (
    (sunrise_sunset_df['sunset_utc'] - sunrise_sunset_df['sunrise_utc'])
    .dt.total_seconds() / 3600
)

print(f"\nDaylight hours at build site (lat={SITE2_LAT}, lon={SITE2_LON})  [M2 buoy: lat={M2_LAT}, lon={M2_LON}]:")
_dl_monthly = (sunrise_sunset_df
               .assign(month=pd.to_datetime(sunrise_sunset_df['date']).dt.month)
               .groupby('month')['daylight_h'].mean())
for m, h in _dl_monthly.items():
    print(f"  {calendar.month_abbr[m]:>3}: {h:.1f} h/day")

# ── Tag each hourly record as daylight or not ─────────────────
wave_df = wave_df.merge(sunrise_sunset_df[['date', 'sunrise_utc', 'sunset_utc']],
                        on='date', how='left')
wave_df['is_daylight'] = (
    (wave_df['time'] >= wave_df['sunrise_utc']) &
    (wave_df['time'] <  wave_df['sunset_utc'])
)


def _hours_in_qualifying_blocks(ok_arr, min_block_h):
    """
    Return a boolean array of the same length as ok_arr where True indicates
    that hour is part of a run of >= min_block_h consecutive True values.
    """
    result = np.zeros(len(ok_arr), dtype=bool)
    n = len(ok_arr)
    i = 0
    while i < n:
        if ok_arr[i]:
            j = i + 1
            while j < n and ok_arr[j]:
                j += 1
            if (j - i) >= min_block_h:
                result[i:j] = True
            i = j
        else:
            i += 1
    return result


def _window_stats(hourly_df, in_block_col, vessel, cfg):
    """Build annual/monthly summary from per-hour in-qualifying-block flags."""
    annual = (hourly_df.groupby('year')[in_block_col]
              .agg(operable_hours='sum', total_hours='count').reset_index())
    annual['pct'] = 100.0 * annual['operable_hours'] / annual['total_hours']
    mean_h = annual['operable_hours'].mean()
    mean_p = annual['pct'].mean()
    monthly = (hourly_df.groupby('month')[in_block_col].mean() * 100).reset_index()
    monthly.columns = ['month', 'pct_operable']
    monthly['month_name'] = monthly['month'].apply(lambda m: calendar.month_abbr[m])
    return {
        'annual':            annual,
        'monthly':           monthly,
        'mean_annual_hours': mean_h,
        'mean_annual_pct':   mean_p,
        'max_hs':            cfg['max_hs'],
        'color':             cfg['color'],
    }


def _compute_window_accessibility(hourly_tbl, daylight_only, label, vessel_windows=None):
    """
    For each vessel and each calendar day, check whether there is a
    stretch of >= vessel window_h consecutive hours where the condition
    is satisfied.  Three masks are evaluated:
      * combined  : wind <= limit  AND  Hs <= limit
      * wave_only : Hs <= limit  (ignoring wind)
      * wind_only : wind <= limit  (ignoring wave)

    If daylight_only=True, only hours between sunrise and sunset are
    considered.  Restricting to fewer hours makes it harder to find a
    qualifying window, so daylight-only correctly gives <= operable days
    compared with the 24-h analysis.
    vessel_windows: optional dict mapping vessel name -> min consecutive hours.
                    Defaults to MIN_WINDOW_HOURS for any vessel not in the dict.
    """
    if daylight_only:
        h = hourly_tbl[hourly_tbl['is_daylight']].copy()
    else:
        h = hourly_tbl.copy()

    # Base daily skeleton (all calendar dates present in data)
    daily = (hourly_tbl
             .groupby(['date', 'year', 'month'])
             .size().rename('n_obs').reset_index())

    combined_res  = {}
    wave_only_res = {}
    wind_only_res = {}

    for vessel, cfg in VESSELS.items():
        # Jack-up Barge operates 24/7 (weeks on-site) — never restrict to daylight hours
        _h = hourly_tbl.copy() if (daylight_only and vessel == 'Jack-up Barge') else h
        min_w = vessel_windows.get(vessel, MIN_WINDOW_HOURS) if vessel_windows else MIN_WINDOW_HOURS
        ok_comb = ((_h['wind_ms'] <= cfg['max_wind_ms']) &
                   (_h['hs']      <= cfg['max_hs'])).fillna(False)
        ok_wave = (_h['hs']      <= cfg['max_hs']).fillna(False)
        ok_wind = (_h['wind_ms'] <= cfg['max_wind_ms']).fillna(False)

        for ok_mask, res_dict, tag in [
            (ok_comb, combined_res,  'comb'),
            (ok_wave, wave_only_res, 'wave'),
            (ok_wind, wind_only_res, 'wind'),
        ]:
            _hh = _h.copy()
            # Mark each hour that belongs to a qualifying block of >= min_w consecutive hours
            in_block = _hours_in_qualifying_blocks(ok_mask.values, min_w)
            _hh['_inblock'] = in_block
            res_dict[vessel] = _window_stats(_hh, '_inblock', vessel, cfg)

            # Derive per-day operable flag for calendar heatmap
            if tag == 'comb':
                _day_op = (_hh.groupby('date')['_inblock'].any()
                           .reset_index()
                           .rename(columns={'_inblock': f'op_{vessel}'}))
                daily = daily.merge(_day_op, on='date', how='left')
                daily[f'op_{vessel}'] = daily[f'op_{vessel}'].fillna(False)

        r_c = combined_res[vessel]
        r_w = wave_only_res[vessel]
        r_n = wind_only_res[vessel]
        _vlabel = 'all hours (24/7 – on-site campaign)' if (daylight_only and vessel == 'Jack-up Barge') else label
        print(f"\n{vessel} (Hs <= {cfg['max_hs']} m, wind <= {cfg['max_wind_ms']} m/s  [{_vlabel}])")
        print(f"  Min qualifying block          : {min_w} h")
        print(f"  Wave-only operable time       : {r_w['mean_annual_hours']:.0f} h/yr  ({r_w['mean_annual_pct']:.1f} %)")
        print(f"  Wind-only operable time       : {r_n['mean_annual_hours']:.0f} h/yr  ({r_n['mean_annual_pct']:.1f} %)")
        print(f"  Combined  operable time       : {r_c['mean_annual_hours']:.0f} h/yr  ({r_c['mean_annual_pct']:.1f} %)")
        print(r_c['annual'][['year', 'operable_hours', 'total_hours', 'pct']].to_string(index=False))

    # ── Augment daily with Hs/daylight statistics ────────────
    _hs_daily = (hourly_tbl.groupby('date')['hs']
                 .agg(hs_max='max', hs_mean='mean').reset_index())
    daily = daily.merge(_hs_daily, on='date', how='left')

    if daylight_only:
        _dl_hrs = hourly_tbl[hourly_tbl['is_daylight']]
        _hs_dl  = (_dl_hrs.groupby('date')['hs']
                   .agg(hs_max_dl='max', hs_mean_dl='mean').reset_index())
        _dl_obs = (_dl_hrs.groupby('date').size()
                   .rename('daylight_obs').reset_index())
        _ss = (hourly_tbl[['date', 'sunrise_utc', 'sunset_utc']]
               .drop_duplicates('date'))
        _ss = _ss.copy()
        _ss['daylight_h'] = (
            (_ss['sunset_utc'] - _ss['sunrise_utc']).dt.total_seconds() / 3600
        )
        daily = (daily
                 .merge(_hs_dl,  on='date', how='left')
                 .merge(_dl_obs, on='date', how='left')
                 .merge(_ss,     on='date', how='left'))

    return combined_res, wave_only_res, wind_only_res, daily


# Minimum qualifying block length: CTV = 4 h, Jack-up Barge = 20 h
_vessel_windows = {'Crew Transfer Vessel': 4, 'Jack-up Barge': 20}

print("\n--- ALL HOURS (24 h/day) ---")
q2_res_24h, q2_wave_24h, q2_wind_24h, daily_24h = _compute_window_accessibility(
    wave_df, False, 'all hours', _vessel_windows)

print("\n--- DAYLIGHT HOURS ONLY ---")
q2_res_dl, q2_wave_dl, q2_wind_dl, daily_dl = _compute_window_accessibility(
    wave_df, True, 'daylight only', _vessel_windows)

# ── Figure 1: Hs time-series with vessel thresholds ──────────
fig, ax = plt.subplots(figsize=(14, 4))
ax.plot(wave_df['time'], wave_df['hs'], lw=0.4, color='steelblue', alpha=0.7, label='Hs (hourly)')
for vessel, cfg in VESSELS.items():
    ax.axhline(cfg['max_hs'], ls='--', lw=1.5, color=cfg['color'],
               label=f"{vessel} limit ({cfg['max_hs']} m)")
ax.set_xlabel('Date')
ax.set_ylabel('Significant Wave Height, Hs (m)')
ax.set_title('Significant Wave Height Time-Series with Operational Limits')
ax.legend(fontsize=9)
ax.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_01_Hs_timeseries.png'), dpi=150, bbox_inches='tight')
plt.close()
print("\nSaved: Q2_01_Hs_timeseries.png")

# ── Figure 1b: Daylight hours through the year ───────────────
fig, ax = plt.subplots(figsize=(11, 4))
_dl_plot = sunrise_sunset_df.copy()
_dl_plot['date_dt'] = pd.to_datetime(_dl_plot['date'])
ax.plot(_dl_plot['date_dt'], _dl_plot['daylight_h'],
        lw=0.8, color='#FFA500', alpha=0.8, label='Daily daylight hours')
ax.fill_between(_dl_plot['date_dt'],
                _dl_plot['sunrise_utc'].apply(lambda t: t.hour + t.minute/60),
                _dl_plot['sunset_utc'].apply(lambda t: t.hour + t.minute/60),
                alpha=0.15, color='#FFA500', label='Daylight window (UTC hour)')
ax.set_xlabel('Date')
ax.set_ylabel('Hours')
ax.set_title(f'Daylight Hours at Build Site (lat={SITE2_LAT}\u00b0N, lon={abs(SITE2_LON):.4f}\u00b0W) \u2013 UTC\n(M2 buoy wind data: lat={M2_LAT}\u00b0N, lon={abs(M2_LON):.4f}\u00b0W)')
ax.legend(fontsize=9)
ax.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_01b_daylight_hours.png'), dpi=150, bbox_inches='tight')
plt.close()
print("Saved: Q2_01b_daylight_hours.png")

# ── Figure 2: Hs histogram + Weibull – all hours & daylight ──
_hs_all = wave_df['hs'].dropna()
_hs_dl  = wave_df.loc[wave_df['is_daylight'], 'hs'].dropna()

_hs_k_all, _, _hs_C_all = weibull_min.fit(_hs_all, floc=0)
_hs_k_dl,  _, _hs_C_dl  = weibull_min.fit(_hs_dl,  floc=0)

fig, axes = plt.subplots(1, 2, figsize=(16, 5), sharey=True)
for ax, hs_data, hs_k, hs_C, mode_title in [
    (axes[0], _hs_all, _hs_k_all, _hs_C_all, 'All Hours (24 h/day)'),
    (axes[1], _hs_dl,  _hs_k_dl,  _hs_C_dl,  'Daylight Hours Only'),
]:
    ax.hist(hs_data, bins=80, color='steelblue', alpha=0.65,
            edgecolor='white', density=True, label='Observed Hs')
    _x_hs = np.linspace(0, hs_data.max(), 500)
    ax.plot(_x_hs, weibull_min.pdf(_x_hs, hs_k, loc=0, scale=hs_C),
            'r-', lw=2, label=f'Weibull fit\nk={hs_k:.3f}, C={hs_C:.3f} m')
    ax.axvline(hs_data.mean(), color='black', lw=1.5, ls='-',
               label=f'Mean = {hs_data.mean():.3f} m')
    ax.axvline(hs_data.median(), color='grey', lw=1.2, ls=':',
               label=f'Median = {hs_data.median():.3f} m')
    for vessel, cfg in VESSELS.items():
        ax.axvline(cfg['max_hs'], ls='--', lw=1.5, color=cfg['color'],
                   label=f"{vessel} limit ({cfg['max_hs']} m)")
    ax.set_xlabel('Significant Wave Height, Hs (m)')
    ax.set_ylabel('Probability Density')
    ax.set_title(mode_title, fontsize=11)
    ax.legend(fontsize=8)
    ax.grid(True, alpha=0.3)
fig.suptitle('Distribution of Significant Wave Height with Weibull Fit',
             fontsize=12, fontweight='bold')
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_02_Hs_histogram.png'), dpi=150, bbox_inches='tight')
plt.close()
print("Saved: Q2_02_Hs_histogram.png")

print(f"\nWeibull fit to Hs  (all hours)   : k={_hs_k_all:.4f}, C={_hs_C_all:.4f} m")
print(f"Weibull fit to Hs  (daylight)    : k={_hs_k_dl:.4f},  C={_hs_C_dl:.4f} m")

# ── Figure 3: Monthly mean Hs ─────────────────────────────────
monthly_mean_hs = wave_df.groupby('month')['hs'].mean()
fig, ax = plt.subplots(figsize=(9, 4))
ax.bar(monthly_mean_hs.index, monthly_mean_hs.values, color='steelblue', alpha=0.8)
ax.axhline(mean_hs, color='black', ls='--', lw=1.2,
           label=f'Annual mean = {mean_hs:.3f} m')
ax.set_xticks(range(1, 13))
ax.set_xticklabels([calendar.month_abbr[m] for m in range(1, 13)])
ax.set_xlabel('Month')
ax.set_ylabel('Mean Hs (m)')
ax.set_title('Monthly Mean Significant Wave Height')
ax.legend(fontsize=9)
ax.grid(True, axis='y', alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_03_monthly_mean_Hs.png'), dpi=150, bbox_inches='tight')
plt.close()
print("Saved: Q2_03_monthly_mean_Hs.png")

# ── Figure 4: Monthly accessibility — both analyses ───────────
# Two subplots: left = all hours, right = daylight only
fig, axes = plt.subplots(1, 2, figsize=(16, 5), sharey=True)
x = np.arange(1, 13)
_month_labels = [calendar.month_abbr[m] for m in range(1, 13)]

for ax_sub, (q2_res, mode_title) in zip(
        axes,
        [(q2_res_24h, 'All Hours (24 h/day)'),
         (q2_res_dl,  'Daylight Hours Only')]):
    width = 0.38
    for i, (vessel, res) in enumerate(q2_res.items()):
        offset = (i - 0.5) * width
        ax_sub.bar(x + offset, res['monthly']['pct_operable'], width,
                   color=res['color'], alpha=0.85, label=vessel)
    ax_sub.set_xticks(x)
    ax_sub.set_xticklabels(_month_labels, rotation=45, ha='right', fontsize=8)
    ax_sub.set_xlabel('Month')
    ax_sub.set_title(mode_title, fontsize=11)
    ax_sub.yaxis.set_major_formatter(mticker.PercentFormatter())
    ax_sub.set_ylim(0, 105)
    ax_sub.grid(True, axis='y', alpha=0.3)
    ax_sub.legend(fontsize=9)

axes[0].set_ylabel('Operable hours (%)')
fig.suptitle('Monthly Operational Accessibility by Vessel Type\n'
             '(% of hours within conditions AND part of a \u226512 h qualifying block, averaged 2021\u20132025)',
             fontsize=12, fontweight='bold')
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_04_monthly_accessibility.png'), dpi=150, bbox_inches='tight')
plt.close()
print("Saved: Q2_04_monthly_accessibility.png")

# ── Figure 5: Annual operable days — both analyses ────────────
vessel_keys = list(VESSELS.keys())
years_ref   = q2_res_24h[vessel_keys[0]]['annual']['year'].values
x = np.arange(len(years_ref))

fig, axes = plt.subplots(1, 2, figsize=(16, 5), sharey=True)
for ax_sub, (q2_res, mode_title) in zip(
        axes,
        [(q2_res_24h, 'All Hours (24 h/day)'),
         (q2_res_dl,  'Daylight Hours Only')]):
    width = 0.38
    for i, (vessel, res) in enumerate(q2_res.items()):
        offset = (i - 0.5) * width
        ax_sub.bar(x + offset, res['annual']['operable_hours'], width,
                   color=res['color'], alpha=0.85, label=vessel)
        ax_sub.axhline(res['mean_annual_hours'], color=res['color'], ls='--', lw=1.3,
                       label=f"{vessel} mean = {res['mean_annual_hours']:.0f} h/yr")
    ax_sub.set_xticks(x)
    ax_sub.set_xticklabels(years_ref.astype(str))
    ax_sub.set_xlabel('Year')
    ax_sub.set_title(mode_title, fontsize=11)
    ax_sub.grid(True, axis='y', alpha=0.3)
    handles, labels = ax_sub.get_legend_handles_labels()
    ax_sub.legend(handles, labels, fontsize=8, loc='lower right')

axes[0].set_ylabel('Operable hours per year')
fig.suptitle('Annual Operable Hours by Vessel Type\n'
             '(hours within conditions AND part of a \u226512 h qualifying block)',
             fontsize=12, fontweight='bold')
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_05_annual_operable_days.png'), dpi=150, bbox_inches='tight')
plt.close()
print("Saved: Q2_05_annual_operable_days.png")

# ── Figure 5b: Side-by-side comparison all vs daylight ────────
fig, ax = plt.subplots(figsize=(12, 5))
_n_v = len(VESSELS)
_n_y = len(years_ref)
_w   = 0.18
_offsets_24h = np.array([-0.30, -0.10])
_offsets_dl  = np.array([ 0.10,  0.30])
_hatches     = ['', '//']

for vi, vessel in enumerate(vessel_keys):
    res_24h = q2_res_24h[vessel]
    res_dl  = q2_res_dl[vessel]
    color   = VESSELS[vessel]['color']
    x_pos   = np.arange(_n_y) + _offsets_24h[vi]
    ax.bar(x_pos, res_24h['annual']['operable_hours'], _w,
           color=color, alpha=0.85, hatch='',
           label=f"{vessel} – all hours")
    x_pos2  = np.arange(_n_y) + _offsets_dl[vi]
    ax.bar(x_pos2, res_dl['annual']['operable_hours'], _w,
           color=color, alpha=0.50, hatch='//',
           label=f"{vessel} – daylight only")

ax.set_xticks(np.arange(_n_y))
ax.set_xticklabels(years_ref.astype(str))
ax.set_xlabel('Year')
ax.set_ylabel('Operable hours per year')
ax.set_title('Annual Operable Hours: All Hours vs Daylight Hours Only\n'
             '(hours within conditions AND part of a ≥12 h qualifying block)')
ax.legend(fontsize=8, ncol=2)
ax.grid(True, axis='y', alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_05b_annual_comparison.png'), dpi=150, bbox_inches='tight')
plt.close()
print("Saved: Q2_05b_annual_comparison.png")

# ── Figure 6: Calendar heatmaps (daylight-only operability) ──
cmap_cal    = ListedColormap(['#e74c3c', '#2ecc71'])
green_patch = mpatches.Patch(color='#2ecc71', label='Operable')
red_patch   = mpatches.Patch(color='#e74c3c', label='Not operable')

def _save_calendar_heatmap(daily_tbl, vessel, max_hs, mode_tag, mode_label, min_w=None):
    """
    Render and save a GitHub-style calendar heatmap showing daily operability.

    Parameters
    ----------
    daily_tbl  : DataFrame with columns ['year', 'date', f'op_{vessel}']
    vessel     : vessel name string (must be a key in VESSELS)
    max_hs     : Hs threshold used for this vessel [m]
    mode_tag   : short string appended to the output filename (e.g. '24h', 'daylight')
    mode_label : human-readable label shown in the figure title
    min_w      : minimum qualifying consecutive-hour block length [h]; defaults to MIN_WINDOW_HOURS
    """
    years  = sorted(daily_tbl['year'].unique())
    n_yr   = len(years)
    fig, axes = plt.subplots(1, n_yr, figsize=(4.5 * n_yr, 6), sharey=True)
    if n_yr == 1:
        axes = [axes]
    col = f'op_{vessel}'
    for ax_cal, yr in zip(axes, years):
        yr_sub = daily_tbl[daily_tbl['year'] == yr][['date', col]].copy()
        yr_sub['date'] = pd.to_datetime(yr_sub['date'])
        yr_sub = yr_sub.set_index('date')
        yr_start  = pd.Timestamp(f'{yr}-01-01')
        all_dates = pd.date_range(yr_start, f'{yr}-12-31', freq='D')
        grid = np.full((7, 54), np.nan)
        for d in all_dates:
            c = (d.dayofyear - 1 + yr_start.dayofweek) // 7
            r = d.dayofweek
            if d in yr_sub.index:
                grid[r, c] = float(yr_sub.loc[d, col])
        masked = np.ma.masked_invalid(grid)
        ax_cal.imshow(masked, aspect='auto', cmap=cmap_cal, vmin=0, vmax=1,
                      origin='upper', interpolation='none')
        ax_cal.set_title(str(yr), fontsize=11)
        ax_cal.set_xlabel('Week', fontsize=9)
        ax_cal.set_yticks(range(7))
        ax_cal.set_yticklabels(['Mon','Tue','Wed','Thu','Fri','Sat','Sun'], fontsize=8)
    fig.legend(handles=[green_patch, red_patch], loc='upper right', fontsize=9)
    sname = vessel.replace(' ', '_')
    max_wind = VESSELS[vessel]['max_wind_ms']
    _mw_label = min_w if min_w is not None else MIN_WINDOW_HOURS
    fig.suptitle(f'Calendar Heatmap \u2013 {vessel} [{mode_label}]\n'
                 f'(green = day contains hours within a \u2265{_mw_label} h qualifying block, '
                 f'Hs \u2264 {max_hs} m AND wind \u2264 {max_wind} m/s)',
                 fontsize=11, fontweight='bold')
    plt.tight_layout()
    fn = f'Q2_06_calendar_{sname}_{mode_tag}.png'
    plt.savefig(os.path.join(Q2_DIR, fn), dpi=150, bbox_inches='tight')
    plt.close()
    print(f"Saved: {fn}")

for vessel, cfg in VESSELS.items():
    _mw = _vessel_windows.get(vessel, MIN_WINDOW_HOURS)
    _save_calendar_heatmap(daily_24h, vessel, cfg['max_hs'], '24h', 'All Hours', min_w=_mw)
    # Jack-up Barge operates 24/7 — no daylight-only heatmap
    if vessel != 'Jack-up Barge':
        _save_calendar_heatmap(daily_dl, vessel, cfg['max_hs'], 'daylight', 'Daylight Only', min_w=_mw)

# ── Wind-based operational days (Weibull from Q1) ────────────
print("\n" + "="*60)
print("WIND-BASED OPERATIONAL DAYS (Weibull from Q1)")
print("="*60)

# Wind speed operational limits from assignment brief (applied at REF_HEIGHT = 4 m)
VESSEL_WIND_LIMITS = {
    'Crew Transfer Vessel': 12.0,   # m/s
    'Jack-up Barge':        10.0,   # m/s
}

# Weibull parameters from Q1: shape_k, scale_c fitted at measurement height (4 m)
_C_w = scale_c
_k_w = shape_k
print(f"\nWeibull parameters (from Q1, at {REF_HEIGHT} m measurement height):")
print(f"  C (scale) = {_C_w:.4f} m/s")
print(f"  k (shape) = {_k_w:.4f}")
print(f"\nVessel wind speed limits (at {REF_HEIGHT} m):")
for v, lim in VESSEL_WIND_LIMITS.items():
    print(f"  {v}: {lim} m/s")

# Weibull CDF: P(wind <= v) = 1 - exp(-(v/C)^k)
def _wcdf(v, C, k):
    return 1.0 - np.exp(-(v / C) ** k)

# Mean daylight hours per calendar month from astral (already computed above)
_dl_hrs        = _dl_monthly.to_dict()                              # {month: mean h/day}
_avg_month_days = {m: calendar.monthrange(2023, m)[1] for m in range(1, 13)}

wind_res = {}
for vessel in VESSEL_WIND_LIMITS:
    v_lim  = VESSEL_WIND_LIMITS[vessel]
    color  = VESSELS[vessel]['color']
    p_hr   = _wcdf(v_lim, _C_w, _k_w)

    # ALL HOURS: P(full 24-h day operable) = P_hr^24
    # Assumes consecutive hourly wind speeds are statistically independent —
    # a simplification that over-estimates operability because real wind speed
    # exhibits positive autocorrelation (calm spells persist).  The result
    # should therefore be treated as an upper-bound estimate.
    p_d24  = p_hr ** 24
    ann_24 = sum(_avg_month_days[m] * p_d24 for m in range(1, 13))  # [days/yr]

    # DAYLIGHT ONLY: n_dl varies by month as sunrise/sunset shifts seasonally.
    # Same independence assumption applies; n_dl < 24 so p_hr^n_dl > p_hr^24,
    # giving more operable days in the daylight-only scenario as expected.
    ann_dl = 0.0
    for m in range(1, 13):
        n_dl    = _dl_hrs.get(m, 12.0)          # [h] mean daylight hours for month m
        ann_dl += _avg_month_days[m] * (p_hr ** n_dl)

    wind_res[vessel] = {
        'v_lim':      v_lim,       # [m/s] operational wind speed limit
        'p_hour':     p_hr,        # [–]   P(1 hour within limit)
        'p_day_24h':  p_d24,       # [–]   P(full 24-h day within limit)
        'annual_24h': ann_24,      # [days/yr] expected operable days, all hours
        'pct_24h':    100 * ann_24 / 365,  # [%]
        'annual_dl':  ann_dl,
        'pct_dl':     100 * ann_dl / 365,
        'color':      color,
    }

    print(f"\n{vessel} (wind <= {v_lim} m/s at {REF_HEIGHT} m):")
    print(f"  P(1-hour operable)             : {p_hr:.4f}  ({100*p_hr:.2f} %)")
    print(f"  P(full 24-h day operable)      : {p_d24:.4f}  ({100*p_d24:.2f} %)")
    print(f"  Est. annual days, all hours    : {ann_24:.1f}  ({100*ann_24/365:.1f} %)")
    print(f"  Est. annual days, daylight only: {ann_dl:.1f}  ({100*ann_dl/365:.1f} %)")

print("\n--- COMBINED WIND + WAVE ACCESSIBILITY SUMMARY ---")
for vessel in VESSELS:
    print(f"\n{vessel}:")
    print(f"  24h  : wave-only {q2_wave_24h[vessel]['mean_annual_hours']:.0f} h/yr  "
          f" wind-only {q2_wind_24h[vessel]['mean_annual_hours']:.0f} h/yr  "
          f" combined {q2_res_24h[vessel]['mean_annual_hours']:.0f} h/yr  "
          f"({q2_res_24h[vessel]['mean_annual_pct']:.1f} %)")
    print(f"  DL   : wave-only {q2_wave_dl[vessel]['mean_annual_hours']:.0f} h/yr  "
          f" wind-only {q2_wind_dl[vessel]['mean_annual_hours']:.0f} h/yr  "
          f" combined {q2_res_dl[vessel]['mean_annual_hours']:.0f} h/yr  "
          f"({q2_res_dl[vessel]['mean_annual_pct']:.1f} %)")

# ── Figure 7: Weibull CDF with vessel wind limits ─────────────
fig, ax = plt.subplots(figsize=(10, 5))
_v_range = np.linspace(0, 28, 500)
ax.plot(_v_range, _wcdf(_v_range, _C_w, _k_w) * 100, 'k-', lw=2,
        label=f'Weibull CDF  (k={_k_w:.2f}, C={_C_w:.2f} m/s)')
_ws_sorted = np.sort(df['WindSpeed_ms'].dropna().values)
_ecdf      = np.arange(1, len(_ws_sorted) + 1) / len(_ws_sorted) * 100
ax.plot(_ws_sorted, _ecdf, color='grey', lw=0.8, alpha=0.6,
        label='Empirical CDF (2024 buoy data)')
for vessel, wr in wind_res.items():
    p_pct = wr['p_hour'] * 100
    ax.axvline(wr['v_lim'], ls='--', lw=1.8, color=wr['color'],
               label=f"{vessel} \u2014 {wr['v_lim']} m/s  (P = {p_pct:.1f} %)")
    ax.plot(wr['v_lim'], p_pct, 'o', ms=8, color=wr['color'], zorder=5)
ax.set_xlabel(f'Wind Speed at {REF_HEIGHT} m (m/s)')
ax.set_ylabel('Cumulative Probability (%)')
ax.set_title(f'Weibull CDF \u2013 Wind Speed at {REF_HEIGHT} m with Vessel Operational Limits\n'
             f'(k = {_k_w:.2f}, C = {_C_w:.2f} m/s)')
ax.legend(fontsize=9)
ax.set_xlim(left=0)
ax.set_ylim(0, 102)
ax.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_07_weibull_wind_cdf.png'), dpi=150, bbox_inches='tight')
plt.close()
print("\nSaved: Q2_07_weibull_wind_cdf.png")

# ── Figure 8: Estimated wind-based operable days ──────────────
fig, axes = plt.subplots(1, 2, figsize=(12, 5), sharey=True)
_vlabels = [v.replace(' ', '\n') for v in wind_res]
for ax_sub, (wkey, mode_title) in zip(
        axes,
        [('annual_24h', 'All Hours (24 h/day)'),
         ('annual_dl',  'Daylight Hours Only')]):
    vals = [wind_res[v][wkey] for v in wind_res]
    cols = [wind_res[v]['color'] for v in wind_res]
    bars = ax_sub.bar(_vlabels, vals, color=cols, alpha=0.85, width=0.5)
    for bar, val in zip(bars, vals):
        ax_sub.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 3,
                    f'{val:.0f} d\n({val / 365 * 100:.1f} %)',
                    ha='center', va='bottom', fontsize=9)
    ax_sub.set_title(mode_title, fontsize=11)
    ax_sub.set_ylim(0, 420)
    ax_sub.grid(True, axis='y', alpha=0.3)
axes[0].set_ylabel('Est. operable days / year')
fig.suptitle(f'Estimated Annual Operable Days \u2013 Wind Conditions Only\n'
             f'(Weibull: C = {_C_w:.2f} m/s, k = {_k_w:.2f} at {REF_HEIGHT} m; '
             f'hourly independence assumed)',
             fontsize=11, fontweight='bold')
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_08_wind_operable_days.png'), dpi=150, bbox_inches='tight')
plt.close()
print("Saved: Q2_08_wind_operable_days.png")

# ── Figure 9: Summary — wave-only / wind-only / combined ────────
_vessels_list = list(VESSELS.keys())
_x  = np.arange(len(_vessels_list))
_w3 = 0.22
_cats       = ['Wave only\n(≥12 consec. hrs)',
               'Wind only\n(≥12 consec. hrs)',
               'Combined (both)\n(≥12 consec. hrs)']
_cat_colors = ['#4472C4', '#ED7D31', '#70AD47']

fig, axes = plt.subplots(1, 2, figsize=(14, 6), sharey=True)
for ax_sub, (comb_r, wave_r, wind_r, title) in zip(
        axes,
        [(q2_res_24h, q2_wave_24h, q2_wind_24h, 'All Hours (24 h/day)'),
         (q2_res_dl,  q2_wave_dl,  q2_wind_dl,  'Daylight Hours Only')]):
    wave_v = [wave_r[v]['mean_annual_hours'] for v in _vessels_list]
    wind_v = [wind_r[v]['mean_annual_hours'] for v in _vessels_list]
    comb_v = [comb_r[v]['mean_annual_hours'] for v in _vessels_list]
    for oi, (vals, lbl, col) in enumerate(zip(
            [wave_v, wind_v, comb_v], _cats, _cat_colors)):
        bars = ax_sub.bar(_x + (oi - 1) * _w3, vals, _w3,
                          color=col, alpha=0.85, label=lbl)
        for bar, val in zip(bars, vals):
            ax_sub.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 30,
                        f'{val:.0f}', ha='center', va='bottom', fontsize=7)
    ax_sub.set_xticks(_x)
    ax_sub.set_xticklabels([v.replace(' ', '\n') for v in _vessels_list], fontsize=9)
    ax_sub.set_title(title, fontsize=11)
    ax_sub.legend(fontsize=7, loc='upper right')
    ax_sub.grid(True, axis='y', alpha=0.3)
    ax_sub.set_ylim(0, 9500)
axes[0].set_ylabel('Mean annual operable hours (2021–2025)')
fig.suptitle('Annual Operable Hours: Wave-Only vs Wind-Only vs Combined\n'
             '(actual observed data; hours within conditions AND part of a ≥12 h qualifying block)',
             fontsize=12, fontweight='bold')
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_09_combined_operability.png'), dpi=150, bbox_inches='tight')
plt.close()
print("Saved: Q2_09_combined_operability.png")

# ── Excel summary ─────────────────────────────────────────────
xl_path = os.path.join(Q2_DIR, 'Q2_wave_accessibility.xlsx')
with pd.ExcelWriter(xl_path, engine='openpyxl') as writer:
    # Sheet 1: Hs statistics
    pd.DataFrame({
        'Statistic': ['Mean Hs (m)', 'Median Hs (m)', 'Std Dev (m)',
                      '90th Percentile (m)', 'Maximum Hs (m)'],
        'Value':     [round(mean_hs, 4), round(median_hs, 4),
                      round(std_hs, 4), round(p90_hs, 4), round(max_hs_obs, 4)],
    }).to_excel(writer, sheet_name='Hs_Statistics', index=False)

    # Sheets 2 & 3: Annual and monthly wave accessibility — both modes
    for mode_tag, mode_label, q2_res in [
            ('24h',      'All Hours',     q2_res_24h),
            ('Daylight', 'Daylight Only', q2_res_dl)]:
        rows = []
        for vessel, res in q2_res.items():
            for _, row in res['annual'].iterrows():
                rows.append({
                    'Vessel':            vessel,
                    'Mode':              mode_label,
                    'Year':              row['year'],
                    'Operable Hours':    int(row['operable_hours']),
                    'Total Hours':       int(row['total_hours']),
                    'Accessibility %':   round(row['pct'], 1),
                })
        rows.append({k: '' for k in rows[0]})
        for vessel, res in q2_res.items():
            rows.append({
                'Vessel':            vessel,
                'Mode':              mode_label,
                'Year':              'MEAN (all years)',
                'Operable Hours':    round(res['mean_annual_hours'], 1),
                'Total Hours':       '',
                'Accessibility %':   round(res['mean_annual_pct'], 1),
            })
        pd.DataFrame(rows).to_excel(writer, sheet_name=f'Annual_{mode_tag}', index=False)

        month_rows = []
        for vessel, res in q2_res.items():
            for _, row in res['monthly'].iterrows():
                month_rows.append({
                    'Vessel':          vessel,
                    'Mode':            mode_label,
                    'Month':           row['month_name'],
                    'Accessibility %': round(row['pct_operable'], 1),
                })
        pd.DataFrame(month_rows).to_excel(writer, sheet_name=f'Monthly_{mode_tag}', index=False)

    # Sheet: Daily data — all hours
    daily_out_24h = daily_24h[['date', 'year', 'month', 'hs_max', 'hs_mean']].copy()
    daily_out_24h.columns = ['Date', 'Year', 'Month', 'Max Hs (all)', 'Mean Hs (all)']
    for vessel in VESSELS:
        daily_out_24h[f'Operable: {vessel}'] = daily_24h[f'op_{vessel}'].astype(int)
    daily_out_24h.to_excel(writer, sheet_name='Daily_AllHours', index=False)

    # Sheet: Daily data — daylight hours
    daily_out_dl = daily_dl[['date', 'year', 'month',
                              'sunrise_utc', 'sunset_utc', 'daylight_h',
                              'hs_max_dl', 'hs_mean_dl', 'daylight_obs']].copy()
    daily_out_dl.columns = ['Date', 'Year', 'Month',
                             'Sunrise UTC', 'Sunset UTC', 'Daylight Hours',
                             'Max Hs (daylight)', 'Mean Hs (daylight)', 'Daylight records']
    for vessel in VESSELS:
        daily_out_dl[f'Operable: {vessel}'] = daily_dl[f'op_{vessel}'].astype(int)
    daily_out_dl.to_excel(writer, sheet_name='Daily_Daylight', index=False)

    # Sheet: Wind-based estimate (Weibull) – structured layout
    _ws_ww = writer.book.create_sheet('Wind_Weibull')
    _r2 = 1

    # ── Section 1: Weibull parameters ────────────────────────
    _ws_ww.cell(row=_r2, column=1,
                value=f'WEIBULL PARAMETERS  (Q1 fit, at {REF_HEIGHT} m measurement height)').font = SUB_FONT
    _ws_ww.cell(row=_r2, column=1).fill = SUB_FILL
    _ws_ww.cell(row=_r2, column=2).fill = SUB_FILL
    _r2 += 1
    for _lbl, _val in [
        ('Scale parameter C (m/s)', round(_C_w, 4)),
        ('Shape parameter k',       round(_k_w, 4)),
        ('Reference height (m)',    REF_HEIGHT),
    ]:
        _ws_ww.cell(row=_r2, column=1, value=_lbl)
        _ws_ww.cell(row=_r2, column=2, value=_val)
        _r2 += 1
    _r2 += 1  # blank row

    # ── Section 2: Vessel wind speed limits ──────────────────
    _ws_ww.cell(row=_r2, column=1, value='VESSEL WIND SPEED LIMITS').font = SUB_FONT
    _ws_ww.cell(row=_r2, column=1).fill = SUB_FILL
    _ws_ww.cell(row=_r2, column=2).fill = SUB_FILL
    _r2 += 1
    _ws_ww.cell(row=_r2, column=1, value='Vessel').font = SUB_FONT
    _ws_ww.cell(row=_r2, column=2,
                value=f'Max Wind Speed (m/s) at {REF_HEIGHT} m').font = SUB_FONT
    _r2 += 1
    for _vessel, _lim in VESSEL_WIND_LIMITS.items():
        _ws_ww.cell(row=_r2, column=1, value=_vessel)
        _ws_ww.cell(row=_r2, column=2, value=_lim)
        _r2 += 1
    _r2 += 1  # blank row

    # ── Section 3: Per-vessel probability & annual days table ─
    _ww_hdrs = [
        'Vessel',
        f'Max Wind Speed (m/s) at {REF_HEIGHT} m',
        'P(1-hour operable) (%)',
        'P(full 24-h day operable) (%)',
        'Est. annual days - all hours',
        'Accessibility % - all hours',
        'Est. annual days - daylight only',
        'Accessibility % - daylight only',
    ]
    for _ci, _h in enumerate(_ww_hdrs, 1):
        _ws_ww.cell(row=_r2, column=_ci, value=_h).font = SUB_FONT
        _ws_ww.cell(row=_r2, column=_ci).fill = SUB_FILL
    _r2 += 1
    for _vessel, _wr in wind_res.items():
        for _ci, _val in enumerate([
            _vessel,
            _wr['v_lim'],
            round(100 * _wr['p_hour'],    2),
            round(100 * _wr['p_day_24h'], 2),
            round(_wr['annual_24h'],      1),
            round(_wr['pct_24h'],         1),
            round(_wr['annual_dl'],       1),
            round(_wr['pct_dl'],          1),
        ], 1):
            _ws_ww.cell(row=_r2, column=_ci, value=_val)
        _r2 += 1

    # Column widths
    for _ci, _cw in enumerate([34, 34, 26, 30, 30, 24, 34, 28], 1):
        _ws_ww.column_dimensions[get_column_letter(_ci)].width = _cw

    # Sheet: Combined wind + wave estimate
    comb_rows = []
    for vessel in VESSELS:
        wr = wind_res[vessel]
        comb_rows.append({
            'Vessel':                              vessel,
            'Wave operable h/yr (all hrs)':        round(q2_wave_24h[vessel]['mean_annual_hours'], 1),
            'Wave operable % (all hrs)':           round(q2_wave_24h[vessel]['mean_annual_pct'], 1),
            'Combined operable h/yr (all hrs)':    round(q2_res_24h[vessel]['mean_annual_hours'], 1),
            'Combined % (all hrs)':                round(q2_res_24h[vessel]['mean_annual_pct'], 1),
            'Wave operable h/yr (daylight)':       round(q2_wave_dl[vessel]['mean_annual_hours'], 1),
            'Wave operable % (daylight)':          round(q2_wave_dl[vessel]['mean_annual_pct'], 1),
            'Combined operable h/yr (daylight)':   round(q2_res_dl[vessel]['mean_annual_hours'], 1),
            'Combined % (daylight)':               round(q2_res_dl[vessel]['mean_annual_pct'], 1),
        })
    pd.DataFrame(comb_rows).to_excel(writer, sheet_name='Combined_Estimate', index=False)

print("Saved: Q2_wave_accessibility.xlsx")

print("\n" + "="*60)
print("Question 2 - ANALYSIS COMPLETE")
print("="*60)


######
# Extra marks
######

# ==============================================================
# QUESTION 2 (EXTRA) – Consecutive Weather-Window Analysis
# Round-trip operational windows based on Dublin Port distance
# ==============================================================

import itertools

print("\n" + "="*60)
print("Q2 EXTRA – CONSECUTIVE WEATHER WINDOW ANALYSIS")
print("="*60)

print(f"\nDistances from Dublin Port (Alexandra Wharf):")
print(f"  Build site (SITE2) : {SITE_DIST_KM:.1f} km  /  {SITE_DIST_NM:.1f} nm  "
      f"(lat={SITE2_LAT}, lon={SITE2_LON})")
print(f"  M2 buoy            : {M2_DIST_KM:.1f} km  /  {M2_DIST_NM:.1f} nm  "
      f"(lat={M2_LAT}, lon={M2_LON})")

# ── Round-trip window per vessel ──────────────────────────────
for vessel, vp in VESSEL_PARAMS.items():
    print(f"\n{vessel}:")
    print(f"  Distance to build site  : {SITE_DIST_NM:.1f} nm  ({SITE_DIST_KM:.1f} km)")
    print(f"  Transit speed           : {vp['transit_kts']} kts")
    print(f"  One-way transit         : {vp['one_way_h']:.2f} h")
    print(f"  Round-trip transit      : {2*vp['one_way_h']:.2f} h")
    print(f"  On-site operations      : {vp['onsite_h']:.0f} h")
    _w_used = 24 if vessel == 'Jack-up Barge' else vp['window_h']
    _note   = '  (overridden to 24 h – on-site campaign vessel)' if vessel == 'Jack-up Barge' else ' (rounded up)'
    print(f"  Total window needed     : {vp['rt_h_exact']:.1f} h  => {_w_used} h{_note}")

# ── Load and merge wind + wave for 2024 (overlap period) ─────
print("\nBuilding merged wind+wave dataset for 2024...")

_wind = df[['time', 'WindSpeed_ms']].copy()
_wind['time'] = pd.to_datetime(_wind['time']).dt.tz_localize(None)

_wave = wave_df[['time', 'hs']].copy()
_wave['time'] = pd.to_datetime(_wave['time']).dt.tz_localize(None)

merged_ww = pd.merge(_wind, _wave, on='time', how='inner').sort_values('time').reset_index(drop=True)
merged_ww['WindSpeed_ms'] = pd.to_numeric(merged_ww['WindSpeed_ms'], errors='coerce')
merged_ww['hs']           = pd.to_numeric(merged_ww['hs'],           errors='coerce')

# Reindex to a complete hourly grid for 2024 so data gaps break windows
_full_2024 = pd.DataFrame({'time': pd.date_range('2024-01-01', '2024-12-31 23:00', freq='h')})
merged_ww  = pd.merge(_full_2024, merged_ww, on='time', how='left')

print(f"  Hourly records in 2024 grid : {len(merged_ww)}")
print(f"  Records with both obs.      : {merged_ww[['WindSpeed_ms','hs']].notna().all(axis=1).sum()}")
print(f"  Missing / gap hours         : {merged_ww[['WindSpeed_ms','hs']].isna().any(axis=1).sum()}")

merged_ww['month'] = merged_ww['time'].dt.month
merged_ww['date']  = merged_ww['time'].dt.date

# ── Helper: find weather windows of length >= N ───────────────
def find_weather_windows(ok_series, times, n_hours):
    """
    Given a boolean Series 'ok_series' on an hourly grid and the
    corresponding datetime index, return a list of dicts for every
    contiguous block of True values whose length >= n_hours.
    """
    windows = []
    in_block  = False
    start_idx = 0
    ok = ok_series.values

    for i in range(len(ok)):
        if ok[i] and not in_block:
            in_block  = True
            start_idx = i
        elif (not ok[i]) and in_block:
            length = i - start_idx
            if length >= n_hours:
                windows.append({
                    'start':      times[start_idx],
                    'end':        times[i - 1],
                    'duration_h': length,
                })
            in_block = False
    # handle block that runs to end
    if in_block:
        length = len(ok) - start_idx
        if length >= n_hours:
            windows.append({
                'start':      times[start_idx],
                'end':        times[-1],
                'duration_h': length,
            })
    return windows

# ── Run analysis per vessel ───────────────────────────────────
ww_results = {}
times_arr  = merged_ww['time'].values

for vessel, vp in VESSEL_PARAMS.items():
    # Jack-up Barge stays on-site for weeks/months; the 79 h round-trip window
    # is not meaningful here.  Use 24 h (fully operable day) instead.
    n_req = 24 if vessel == 'Jack-up Barge' else vp['window_h']
    color = vp['color']

    # Combined operability: both wind AND wave within limits
    # NaN treated as NOT operable (gap in data breaks a window)
    ok_combined = (
        (merged_ww['WindSpeed_ms'] <= vp['max_wind_ms']) &
        (merged_ww['hs']           <= vp['max_hs'])
    ).fillna(False)

    ok_wind_only = (merged_ww['WindSpeed_ms'] <= vp['max_wind_ms']).fillna(False)
    ok_wave_only = (merged_ww['hs']           <= vp['max_hs']).fillna(False)

    merged_ww[f'ok_combined_{vessel}'] = ok_combined

    windows = find_weather_windows(ok_combined, times_arr, n_req)
    windows_df = pd.DataFrame(windows)
    if not windows_df.empty:
        windows_df['start']   = pd.to_datetime(windows_df['start'])
        windows_df['end']     = pd.to_datetime(windows_df['end'])
        windows_df['month']   = windows_df['start'].dt.month
        windows_df['label']   = windows_df['start'].dt.strftime('%b %d')

    # Monthly count of windows
    monthly_wins = (windows_df.groupby('month')['duration_h'].count()
                   .reindex(range(1, 13), fill_value=0)
                   if not windows_df.empty
                   else pd.Series(0, index=range(1, 13)))

    # Monthly percentage of hours within qualifying blocks
    _in_block = _hours_in_qualifying_blocks(ok_combined.values, n_req)
    merged_ww['_in_block_tmp'] = _in_block
    _monthly_block = (
        merged_ww.groupby('month')
        .agg(_block_h=('_in_block_tmp', 'sum'), _total_h=('_in_block_tmp', 'count'))
        .reindex(range(1, 13), fill_value=0)
    )
    _monthly_block['pct'] = 100.0 * _monthly_block['_block_h'] / _monthly_block['_total_h'].replace(0, np.nan)
    merged_ww.drop(columns=['_in_block_tmp'], inplace=True)

    # Hourly operable fractions
    pct_h_wind = ok_wind_only.sum() / len(ok_wind_only) * 100
    pct_h_wave = ok_wave_only.sum() / len(ok_wave_only) * 100
    pct_h_comb = ok_combined.sum()  / len(ok_combined)  * 100

    ww_results[vessel] = {
        'n_req':         n_req,
        'windows':       windows_df,
        'monthly_wins':  monthly_wins,
        'monthly_block': _monthly_block,
        'n_windows':     len(windows),
        'pct_h_wind':    pct_h_wind,
        'pct_h_wave':    pct_h_wave,
        'pct_h_comb':    pct_h_comb,
        'color':         color,
    }

    print(f"\n{vessel}  [wind <= {vp['max_wind_ms']} m/s, Hs <= {vp['max_hs']} m, window = {n_req} h]")
    print(f"  Hours where wind OK          : {ok_wind_only.sum()} / {len(ok_wind_only)}  ({pct_h_wind:.1f} %)")
    print(f"  Hours where wave OK          : {ok_wave_only.sum()} / {len(ok_wave_only)}  ({pct_h_wave:.1f} %)")
    print(f"  Hours where both OK          : {ok_combined.sum()} / {len(ok_combined)}  ({pct_h_comb:.1f} %)")
    print(f"  Qualifying windows (>= {n_req} h) : {len(windows)}")
    if windows:
        durs = windows_df['duration_h']
        print(f"  Window durations: min={durs.min()} h, mean={durs.mean():.0f} h, max={durs.max()} h")
        print(f"  Monthly distribution:")
        for m, cnt in monthly_wins.items():
            print(f"    {calendar.month_abbr[m]:>3}: {cnt} window(s)")
    print(f"  Monthly operable hours (within >={n_req} h blocks):")
    for m, row in _monthly_block.iterrows():
        print(f"    {calendar.month_abbr[m]:>3}: {int(row['_block_h']):>4} / {int(row['_total_h']):>4} h  ({row['pct']:.1f} %)")

# ── Figure W1: Combined operability mask – 2024 ───────────────
fig, axes = plt.subplots(len(VESSEL_PARAMS), 1, figsize=(16, 7), sharex=True)
for ax_v, (vessel, vp) in zip(axes, VESSEL_PARAMS.items()):
    ok_col = f'ok_combined_{vessel}'
    # Highlight operable hours as filled band
    _ok = merged_ww[ok_col].astype(float)
    ax_v.fill_between(merged_ww['time'], 0, _ok,
                      color=vp['color'], alpha=0.5, step='post',
                      label=f"Both conditions met")
    ax_v.plot(merged_ww['time'],
              (merged_ww['WindSpeed_ms'] / vp['max_wind_ms']).clip(0, 1.2),
              lw=0.5, color='navy', alpha=0.5, label='Wind speed (norm)')
    ax_v.plot(merged_ww['time'],
              (merged_ww['hs'] / vp['max_hs']).clip(0, 1.2),
              lw=0.5, color='steelblue', alpha=0.5, label='Hs (norm)')
    ax_v.axhline(1.0, color='red', ls='--', lw=0.8, label='Threshold (normalised)')
    ax_v.set_ylabel(vessel.replace(' ', '\n'), fontsize=8)
    ax_v.set_ylim(-0.05, 1.3)
    ax_v.legend(fontsize=7, loc='upper right', ncol=4)
    ax_v.grid(True, axis='x', alpha=0.2)
axes[-1].set_xlabel('Date (2024)')
fig.suptitle('Combined Wind + Wave Operability Mask (2024)\n'
             '(green fill = both conditions simultaneously within limits)',
             fontsize=11, fontweight='bold')
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_W1_operability_mask.png'), dpi=150, bbox_inches='tight')
plt.close()
print("\nSaved: Q2_W1_operability_mask.png")

# ── Figure W2: Weather window duration distributions ─────────
fig, axes = plt.subplots(1, 2, figsize=(13, 5))
for ax_v, (vessel, res) in zip(axes, ww_results.items()):
    windows_df = res['windows']
    n_req      = res['n_req']
    if not windows_df.empty:
        ax_v.hist(windows_df['duration_h'], bins=20, color=res['color'],
                  alpha=0.8, edgecolor='white')
        ax_v.axvline(n_req, color='red', ls='--', lw=1.5,
                     label=f'Required: {n_req} h')
        ax_v.axvline(windows_df['duration_h'].mean(), color='black', ls='-', lw=1.2,
                     label=f"Mean: {windows_df['duration_h'].mean():.0f} h")
        ax_v.legend(fontsize=9)
    else:
        ax_v.text(0.5, 0.5, 'No qualifying windows', ha='center', va='center',
                  transform=ax_v.transAxes, fontsize=12)
    ax_v.set_xlabel('Consecutive operable hours')
    ax_v.set_ylabel('Frequency')
    ax_v.set_title(f'{vessel}\n(wind + wave, window >= {n_req} h, 2024)')
    ax_v.grid(True, alpha=0.3)
fig.suptitle('Distribution of Qualifying Weather Window Durations (2024)',
             fontsize=11, fontweight='bold')
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_W2_window_durations.png'), dpi=150, bbox_inches='tight')
plt.close()
print("Saved: Q2_W2_window_durations.png")

# ── Figure W3: Monthly count of qualifying windows ────────────
fig, ax = plt.subplots(figsize=(11, 5))
x  = np.arange(1, 13)
_w = 0.38
for i, (vessel, res) in enumerate(ww_results.items()):
    offset = (i - 0.5) * _w
    ax.bar(x + offset, res['monthly_wins'].values, _w,
           color=res['color'], alpha=0.85, label=vessel)
ax.set_xticks(x)
ax.set_xticklabels([calendar.month_abbr[m] for m in range(1, 13)])
ax.set_xlabel('Month')
ax.set_ylabel('Number of qualifying windows')
ax.set_title('Monthly Count of Qualifying Weather Windows (2024)\n'
             '(wind + wave simultaneously within limits for full round-trip window)',
             fontsize=10, fontweight='bold')
ax.legend(fontsize=9)
ax.grid(True, axis='y', alpha=0.3)
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_W3_monthly_windows.png'), dpi=150, bbox_inches='tight')
plt.close()
print("Saved: Q2_W3_monthly_windows.png")

# ── Figure W4: Gantt-style window timeline ────────────────────
fig, axes = plt.subplots(len(VESSEL_PARAMS), 1, figsize=(15, 6), sharex=True)
for ax_g, (vessel, res) in zip(axes, ww_results.items()):
    windows_df = res['windows']
    n_req      = res['n_req']
    if not windows_df.empty:
        for _, row in windows_df.iterrows():
            dur = row['duration_h']
            bar_color = res['color'] if dur >= n_req else 'lightgrey'
            ax_g.barh(0, left=row['start'], width=pd.Timedelta(hours=dur),
                      height=0.6, color=bar_color, alpha=0.8, edgecolor='white', lw=0.3)
    ax_g.set_yticks([])
    ax_g.set_ylabel(vessel.replace(' ', '\n'), fontsize=8, rotation=0, ha='right', va='center')
    ax_g.set_ylim(-0.5, 0.5)
    ax_g.grid(True, axis='x', alpha=0.2)
    ax_g.set_title(f'{vessel}  |  {res["n_windows"]} qualifying windows >= {n_req} h  '
                   f'(combined wind + wave, 2024)', fontsize=9)
axes[-1].set_xlabel('Date (2024)')
fig.suptitle('Qualifying Weather Windows – Timeline View (2024)',
             fontsize=11, fontweight='bold')
plt.tight_layout()
plt.savefig(os.path.join(Q2_DIR, 'Q2_W4_window_timeline.png'), dpi=150, bbox_inches='tight')
plt.close()
print("Saved: Q2_W4_window_timeline.png")

# ── Excel: weather window results ─────────────────────────────
xl_ww = os.path.join(Q2_DIR, 'Q2_weather_windows.xlsx')
with pd.ExcelWriter(xl_ww, engine='openpyxl') as writer:
    # Summary sheet
    summary_rows = []
    for vessel, vp in VESSEL_PARAMS.items():
        res = ww_results[vessel]
        summary_rows.append({
            'Vessel':                     vessel,
            'Max wind speed (m/s)':       vp['max_wind_ms'],
            'Max Hs (m)':                 vp['max_hs'],
            'Transit speed (kts)':        vp['transit_kts'],
            'One-way transit (h)':        round(vp['one_way_h'], 2),
            'On-site time (h)':           vp['onsite_h'],
            'Round-trip window (h)':      vp['window_h'],
            'Build site lat':             SITE2_LAT,
            'Build site lon':             SITE2_LON,
            'Build site dist from Dublin (nm)': round(SITE_DIST_NM, 1),
            'Build site dist from Dublin (km)': round(SITE_DIST_KM, 1),
            'M2 buoy lat':                M2_LAT,
            'M2 buoy lon':                M2_LON,
            'M2 buoy dist from Dublin (nm)': round(M2_DIST_NM, 1),
            'M2 buoy dist from Dublin (km)': round(M2_DIST_KM, 1),
            'Hours wind OK (%)':          round(res['pct_h_wind'], 1),
            'Hours wave OK (%)':          round(res['pct_h_wave'], 1),
            'Hours both OK (%)':          round(res['pct_h_comb'], 1),
            'Qualifying windows (2024)':  res['n_windows'],
        })
    pd.DataFrame(summary_rows).to_excel(writer, sheet_name='Summary', index=False)

    # Per-vessel window lists
    for vessel, res in ww_results.items():
        sname = vessel[:15].replace(' ', '_')
        if not res['windows'].empty:
            out = res['windows'][['start', 'end', 'duration_h', 'month']].copy()
            out.columns = ['Window Start', 'Window End', 'Duration (h)', 'Month']
            out['Month Name'] = out['Month'].apply(lambda m: calendar.month_abbr[m])
        else:
            out = pd.DataFrame(columns=['Window Start', 'Window End',
                                        'Duration (h)', 'Month', 'Month Name'])
        out.to_excel(writer, sheet_name=sname, index=False)

    # Monthly summary
    month_rows = []
    for vessel, res in ww_results.items():
        for m, cnt in res['monthly_wins'].items():
            month_rows.append({
                'Vessel':             vessel,
                'Month':              calendar.month_abbr[m],
                'Qualifying Windows': int(cnt),
            })
    pd.DataFrame(month_rows).to_excel(writer, sheet_name='Monthly_Summary', index=False)

print("Saved: Q2_weather_windows.xlsx")

print("\n" + "="*60)
print("Q2 EXTRA - CONSECUTIVE WINDOW ANALYSIS COMPLETE")
print("="*60)

